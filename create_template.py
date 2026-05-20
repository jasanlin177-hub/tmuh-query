from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "查詢名單"

headers    = ["姓名", "醫院", "身分證字號", "出生日期(民國年/月/日)\n萬芳醫院可空白", "查詢結果", "查詢時間"]
col_widths = [12, 14, 16, 24, 40, 20]

thin   = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
h_font = Font(name="Arial", bold=True, color="FFFFFF")
h_fill = PatternFill("solid", start_color="2F5496")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font  = h_font
    cell.fill  = h_fill
    cell.alignment = center
    cell.border = border
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 30

# 醫院欄下拉選單
dv = DataValidation(
    type="list",
    formula1='"北醫附醫,萬芳醫院,臺北市立聯合醫院"',
    allow_blank=False,
    showDropDown=False,
)
ws.add_data_validation(dv)
dv.add("B2:B1000")

sample_data = [
    ("王小明", "北醫附醫", "A226127572", "074/06/23"),
    ("李小花", "北醫附醫", "G120417991", "061/08/31"),
    ("方大同", "萬芳醫院", "A123456789", ""),
]

alt_fill = PatternFill("solid", start_color="EEF2F7")
d_font   = Font(name="Arial", size=11)

for r, row in enumerate(sample_data, 2):
    fill = alt_fill if r % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
    for col, val in enumerate(list(row) + ["", ""], 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font   = d_font
        cell.fill   = fill
        cell.border = border
        cell.alignment = Alignment(
            vertical="center",
            horizontal="center" if col != 5 else "left"
        )

ws.freeze_panes = "A2"
wb.save("查詢名單.xlsx")
print("已建立 查詢名單.xlsx（含醫院下拉選單）")
