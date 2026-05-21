import io
import random
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from hospitals import REGISTRY
from hospitals.tmuh import parse_birth_date

st.set_page_config(page_title="掛號查詢系統", layout="centered")

st.markdown("""
<style>
/* ── 全域底色 ── */
[data-testid="stAppViewContainer"] {
    background: linear-gradient(160deg, #f0f4f8 0%, #e8eef5 100%);
}
[data-testid="stHeader"] { background: transparent; }

/* ── 主標題 ── */
h1 {
    background: linear-gradient(135deg, #1a3a6b 0%, #2d6bb5 60%, #4a9fd4 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    font-weight: 800;
    letter-spacing: -0.02em;
    text-shadow: none;
}

/* ── 按鈕 ── */
.stButton > button {
    background: linear-gradient(165deg, #3a70c2 0%, #1e4d99 50%, #163a7a 100%);
    color: #ffffff;
    border: 1px solid #5588d4;
    border-bottom-color: #0d2a5e;
    border-radius: 8px;
    padding: 0.55rem 1.6rem;
    font-weight: 700;
    font-size: 0.95rem;
    letter-spacing: 0.04em;
    box-shadow:
        0 5px 12px rgba(20, 50, 120, 0.35),
        0 2px 4px  rgba(0, 0, 0, 0.25),
        inset 0 1px 0 rgba(255, 255, 255, 0.22),
        inset 0 -2px 0 rgba(0, 0, 0, 0.18);
    transition: all 0.14s ease;
    text-shadow: 0 1px 2px rgba(0,0,0,0.4);
}
.stButton > button:hover {
    background: linear-gradient(165deg, #4d87d9 0%, #2d60b8 50%, #1a4b8c 100%);
    border-color: #7aaee0;
    box-shadow:
        0 8px 20px rgba(20, 50, 120, 0.45),
        0 3px 6px  rgba(0, 0, 0, 0.3),
        inset 0 1px 0 rgba(255, 255, 255, 0.28),
        inset 0 -2px 0 rgba(0, 0, 0, 0.15);
    transform: translateY(-2px);
}
.stButton > button:active {
    background: linear-gradient(165deg, #163a7a 0%, #0d2a5e 100%);
    box-shadow:
        0 2px 6px rgba(20, 50, 120, 0.4),
        inset 0 3px 6px rgba(0, 0, 0, 0.25);
    transform: translateY(1px);
}

/* ── 文字輸入框 ── */
.stTextInput > div > div > input {
    background: linear-gradient(180deg, #ffffff 0%, #f5f8fc 100%);
    border: 1px solid #b8cde0;
    border-radius: 7px;
    box-shadow:
        inset 0 2px 5px rgba(0, 0, 40, 0.08),
        inset 0 1px 2px rgba(0, 0, 40, 0.05),
        0 1px 0 rgba(255, 255, 255, 0.9);
    padding: 0.45rem 0.75rem;
    font-size: 0.95rem;
    color: #1a2a3a;
    transition: all 0.15s ease;
}
.stTextInput > div > div > input:focus {
    border-color: #3a70c2;
    box-shadow:
        inset 0 2px 5px rgba(0, 0, 40, 0.06),
        0 0 0 3px rgba(58, 112, 194, 0.18),
        0 1px 0 rgba(255, 255, 255, 0.9);
    outline: none;
}

/* ── 下拉選單 ── */
.stSelectbox > div > div {
    background: linear-gradient(180deg, #ffffff 0%, #f5f8fc 100%);
    border: 1px solid #b8cde0 !important;
    border-radius: 7px;
    box-shadow:
        inset 0 2px 4px rgba(0, 0, 40, 0.07),
        0 1px 0 rgba(255, 255, 255, 0.9);
}

/* ── Tabs ── */
[data-baseweb="tab-list"] {
    background: linear-gradient(180deg, #dce8f5 0%, #c8daf0 100%);
    border-radius: 10px 10px 0 0;
    padding: 4px 4px 0 4px;
    gap: 4px;
    box-shadow: inset 0 -2px 0 rgba(0,0,0,0.08);
}
[data-baseweb="tab"] {
    background: transparent;
    border-radius: 8px 8px 0 0;
    font-weight: 600;
    color: #3a5a8a;
    padding: 0.5rem 1.2rem;
    border: none;
    transition: all 0.12s ease;
}
[aria-selected="true"][data-baseweb="tab"] {
    background: linear-gradient(180deg, #ffffff 0%, #f0f5fb 100%);
    color: #1a3a6b;
    box-shadow:
        0 -1px 0 #5588d4,
        2px 0 6px rgba(0,0,0,0.08),
        -2px 0 6px rgba(0,0,0,0.08);
}

/* ── 卡片區塊（Tab 內容） ── */
[data-baseweb="tab-panel"] {
    background: linear-gradient(180deg, #f8fbff 0%, #f0f4fa 100%);
    border-radius: 0 0 10px 10px;
    border: 1px solid #c8daf0;
    border-top: none;
    padding: 1.2rem 1.5rem 1.5rem;
    box-shadow:
        0 6px 20px rgba(20, 50, 120, 0.1),
        0 2px 6px rgba(0, 0, 0, 0.07);
}

/* ── Tab 選中底線（覆蓋預設橘紅） ── */
[data-baseweb="tab-highlight"] {
    background-color: #2d60b8 !important;
}

/* ── 檔案上傳 Browse files 按鈕 ── */
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] button {
    background: linear-gradient(165deg, #3a70c2 0%, #1e4d99 50%, #163a7a 100%) !important;
    color: #ffffff !important;
    border: 1px solid #5588d4 !important;
    border-bottom-color: #0d2a5e !important;
    border-radius: 8px !important;
    font-weight: 700 !important;
    letter-spacing: 0.04em !important;
    box-shadow:
        0 5px 12px rgba(20, 50, 120, 0.35),
        0 2px 4px  rgba(0, 0, 0, 0.25),
        inset 0 1px 0 rgba(255, 255, 255, 0.22),
        inset 0 -2px 0 rgba(0, 0, 0, 0.18) !important;
    text-shadow: 0 1px 2px rgba(0,0,0,0.4) !important;
    transition: all 0.14s ease !important;
}
[data-testid="stFileUploaderDropzone"] button:hover,
[data-testid="stFileUploader"] button:hover {
    background: linear-gradient(165deg, #4d87d9 0%, #2d60b8 50%, #1a4b8c 100%) !important;
    box-shadow:
        0 8px 20px rgba(20, 50, 120, 0.45),
        0 3px 6px  rgba(0, 0, 0, 0.3),
        inset 0 1px 0 rgba(255, 255, 255, 0.28),
        inset 0 -2px 0 rgba(0, 0, 0, 0.15) !important;
    transform: translateY(-2px) !important;
}

/* ── 進度條 ── */
[data-testid="stProgress"] > div > div > div > div {
    background: linear-gradient(90deg, #2d60b8, #4a9fd4);
    border-radius: 4px;
}
</style>
""", unsafe_allow_html=True)

st.title("掛號查詢系統")
st.caption("支援：" + "・".join(h.display_name for h in REGISTRY.values()))

_NO_DATA_KEYS = ["查無", "查不到", "無資料", "略過", "需填", "格式錯誤", "伺服器訊息", "無法自動解析"]
_ERROR_KEYS   = ["錯誤", "失敗", "超過", "網路", "逾時"]

def _has_appointment(result: str) -> bool:
    if not result:
        return False
    return not any(k in result for k in _NO_DATA_KEYS + _ERROR_KEYS)


def build_result_xlsx(rows):
    """rows: [(姓名, 醫院, 身分證, 出生日期, 結果, 時間), ...]"""
    wb = Workbook()
    ws = wb.active
    ws.title = "查詢結果"

    headers    = ["姓名", "醫院", "身分證字號", "出生日期", "查詢結果", "查詢時間"]
    col_widths = [12, 18, 16, 16, 45, 20]

    thin   = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(name="Arial", bold=True, color="FFFFFF")
    h_fill = PatternFill("solid", start_color="2F5496")

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ok_fill   = PatternFill("solid", start_color="E2EFDA")
    err_fill  = PatternFill("solid", start_color="FCE4D6")
    alt_fill  = PatternFill("solid", start_color="EEF2F7")
    appt_fill = PatternFill("solid", start_color="FFE0CC")
    d_font    = Font(name="Arial", size=11)
    appt_font = Font(name="Arial", size=11, bold=True, color="C00000")

    for r, row in enumerate(rows, 2):
        result  = row[4] or ""
        has_apt = _has_appointment(result)
        is_err  = any(k in result for k in _ERROR_KEYS)
        if has_apt:
            fill = appt_fill
            font = appt_font
        elif is_err:
            fill = err_fill
            font = d_font
        else:
            fill = ok_fill if r % 2 == 0 else alt_fill
            font = d_font
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font   = font
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left" if col == 5 else "center",
                wrap_text=(col == 5)
            )

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── 單筆查詢 Tab ───────────────────────────────────────────────────────────────
tab_single, tab_batch = st.tabs(["單筆查詢", "批次查詢（XLSX）"])

with tab_single:
    st.subheader("單筆查詢")
    hosp_name = st.selectbox("醫院", list(REGISTRY.keys()), key="single_hosp")
    hosp      = REGISTRY[hosp_name]

    c1, c2 = st.columns(2)
    with c1:
        single_id = st.text_input("身分證字號", placeholder="A123456789",
                                   key="single_id").strip().upper()
    with c2:
        if hosp.needs_birth:
            single_birth = st.text_input("出生日期（民國年/月/日）",
                                          placeholder="074/06/23",
                                          key="single_birth").strip()
        else:
            st.caption(f"{hosp_name} 僅需身分證字號")
            single_birth = ""

    if st.button("查詢", key="btn_single"):
        if not single_id:
            st.warning("請填入身分證字號")
        elif hosp.needs_birth and not single_birth:
            st.warning("請填入出生日期")
        else:
            birth_args = {}
            if hosp.needs_birth:
                parsed = parse_birth_date(single_birth)
                if not parsed:
                    st.error("出生日期格式錯誤，請輸入如 074/06/23")
                    st.stop()
                birth_args = dict(birth_year=parsed[0],
                                  birth_month=parsed[1],
                                  birth_day=parsed[2])
            with st.spinner("查詢中，請稍候..."):
                session = hosp.make_session()
                result  = hosp.query_one(session, single_id, **birth_args)
            if any(k in result for k in ["查不到", "查無"]):
                st.info(result)
            elif any(k in result for k in _ERROR_KEYS):
                st.error(result)
            elif _has_appointment(result):
                st.warning(f"⚠️ 查到掛號資料！")
                st.text(result)
            else:
                st.success("查詢完成")
                st.text(result)

# ── 批次查詢 ───────────────────────────────────────────────────────────────────
with tab_batch:
    st.subheader("批次查詢")
    hosp_list = "、".join(REGISTRY.keys())
    st.caption(
        f"XLSX 欄位順序：**姓名**（可空）、**醫院**（{hosp_list}）、"
        "**身分證字號**、**出生日期**（需出生日期的醫院必填，其餘可空白）"
    )

    uploaded  = st.file_uploader("上傳查詢名單 XLSX", type=["xlsx"])
    query_all = st.checkbox("查詢所有醫院（忽略醫院欄位，同時查詢全部醫院）",
                            value=True)

    if uploaded:
        df = pd.read_excel(uploaded, dtype=str).fillna("")
        df.columns = df.columns.str.strip()
        st.dataframe(df, use_container_width=True)

        if st.button("開始批次查詢", key="btn_batch"):
            sessions = {name: h.make_session() for name, h in REGISTRY.items()}
            rows     = []
            total    = len(df)
            progress = st.progress(0, text="準備中...")
            status   = st.empty()

            for i, row in df.iterrows():
                vals    = list(row.values)
                name    = str(vals[0]).strip() if len(vals) > 0 else ""
                hospital= str(vals[1]).strip() if len(vals) > 1 else ""
                id_no   = str(vals[2]).strip() if len(vals) > 2 else ""
                birth_b = str(vals[3]).strip() if len(vals) > 3 else ""

                progress.progress(i / total, text=f"查詢 {i+1}/{total}：{id_no}")

                if not id_no:
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    rows.append((name, hospital, id_no, birth_b, "身分證字號為空，略過", ts))
                    continue

                if query_all:
                    parts = {}
                    status.text(f"並行查詢 {len(REGISTRY)} 家醫院：{name} {id_no}")
                    parsed = parse_birth_date(birth_b) if birth_b else None
                    with ThreadPoolExecutor(max_workers=len(REGISTRY)) as ex:
                        futures = {}
                        for hname, hosp in REGISTRY.items():
                            if hosp.needs_birth:
                                if not birth_b:
                                    parts[hname] = "需填出生日期，略過"
                                    continue
                                if not parsed:
                                    parts[hname] = "出生日期格式錯誤"
                                    continue
                                futures[ex.submit(
                                    hosp.query_one, sessions[hname], id_no,
                                    parsed[0], parsed[1], parsed[2]
                                )] = hname
                            else:
                                futures[ex.submit(
                                    hosp.query_one, sessions[hname], id_no
                                )] = hname
                        for future in as_completed(futures):
                            hname = futures[future]
                            try:
                                parts[hname] = future.result()
                            except Exception as e:
                                parts[hname] = f"查詢失敗：{e}"
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    for hname in REGISTRY:
                        if hname in parts:
                            rows.append((name, hname, id_no, birth_b, parts[hname], ts))
                else:
                    hosp_tag = hospital if hospital else "北醫附醫"
                    hosp     = REGISTRY.get(hosp_tag, REGISTRY["北醫附醫"])
                    status.text(f"[{hosp_tag}] {name} {id_no}")
                    if hosp.needs_birth:
                        if not birth_b:
                            result = f"{hosp_tag}需填出生日期"
                        else:
                            parsed = parse_birth_date(birth_b)
                            if not parsed:
                                result = "出生日期格式錯誤"
                            else:
                                result = hosp.query_one(sessions[hosp_tag], id_no,
                                                        birth_year=parsed[0],
                                                        birth_month=parsed[1],
                                                        birth_day=parsed[2])
                    else:
                        result = hosp.query_one(sessions[hosp_tag], id_no)
                    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    rows.append((name, hosp_tag, id_no, birth_b, result, ts))

                if i < total - 1:
                    time.sleep(random.uniform(0.3, 0.8))

            progress.progress(1.0, text="查詢完成！")
            status.empty()

            result_df = pd.DataFrame(
                rows,
                columns=["姓名", "醫院", "身分證字號", "出生日期", "查詢結果", "查詢時間"]
            )
            st.dataframe(result_df, use_container_width=True)

            appt_hits = [(r[0], r[2], r[1], r[4]) for r in rows if _has_appointment(r[4])]
            if appt_hits:
                st.warning(f"⚠️ 發現 {len(appt_hits)} 筆有掛號資料（粗體紅字標示於 XLSX）：")
                for a_name, a_id, a_hosp, a_result in appt_hits:
                    label = f"{a_name}（{a_id}）" if a_name else a_id
                    with st.expander(f"🔴 {label} ── {a_hosp}", expanded=True):
                        st.text(a_result)

            xlsx_bytes = build_result_xlsx(rows)
            fname = f"查詢結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="⬇️ 下載結果 XLSX",
                data=xlsx_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
