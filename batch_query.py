"""
批次掛號查詢 - 從 XLSX 讀取名單，跨醫院並行查詢後將結果寫回新檔案
支援醫院：依 hospitals/REGISTRY 自動支援所有已登記醫院
用法: python batch_query.py [查詢名單.xlsx]

XLSX 欄位順序：姓名 | 醫院 | 身分證字號 | 出生日期(民國年/月/日) | 查詢結果 | 查詢時間
並行策略：不同醫院同時查（各打各的伺服器），同一醫院依序查並加 0.3~0.8s 隨機延遲
"""

import sys
import time
import random
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import urllib3
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from hospitals import REGISTRY
from hospitals.tmuh import parse_birth_date

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

DELAY_MIN = 0.3
DELAY_MAX = 0.8

_print_lock = threading.Lock()


def _log(msg):
    with _print_lock:
        print(msg, flush=True)


def _query_group(hospital_tag, rows_data, session):
    """查詢同一醫院的所有列，回傳 {row_num: (result, timestamp)} dict"""
    hosp = REGISTRY.get(hospital_tag, REGISTRY["北醫附醫"])
    results = {}

    for i, (row_num, id_no, birth_raw) in enumerate(rows_data):
        _log(f"  [{hosp.display_name}] {id_no} ({birth_raw})...")

        if hosp.needs_birth:
            if not birth_raw:
                result = f"{hosp.display_name}需填出生日期"
            else:
                parsed = parse_birth_date(birth_raw)
                if not parsed:
                    result = "出生日期格式錯誤"
                else:
                    y, m, d = parsed
                    result = hosp.query_one(session, id_no, y, m, d)
        else:
            result = hosp.query_one(session, id_no)

        _log(f"    → {result[:60]}")
        results[row_num] = (result, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        if i < len(rows_data) - 1:
            delay = random.uniform(DELAY_MIN, DELAY_MAX)
            time.sleep(delay)

    return results


def run_batch(xlsx_path):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    result_col = 5
    time_col   = 6

    ok_fill  = PatternFill("solid", start_color="E2EFDA")
    err_fill = PatternFill("solid", start_color="FCE4D6")
    font     = Font(name="Arial", size=11)

    # 讀取所有有效列，按醫院分組
    groups = defaultdict(list)
    for row in ws.iter_rows(min_row=2):
        if len(row) < 4:
            continue
        id_no    = str(row[2].value or "").strip()
        hospital = str(row[1].value or "").strip() or "北醫附醫"
        if not id_no:
            continue
        birth_raw = str(row[3].value or "").strip()
        groups[hospital].append((row[0].row, id_no, birth_raw))

    if not groups:
        print("無有效資料列")
        return

    hospital_list = sorted(groups)
    print(f"共 {sum(len(v) for v in groups.values())} 筆，"
          f"分 {len(groups)} 家醫院並行查詢：{hospital_list}")

    sessions = {tag: REGISTRY.get(tag, REGISTRY["北醫附醫"]).make_session()
                for tag in groups}

    # 跨醫院並行，同醫院內依序
    all_results = {}
    with ThreadPoolExecutor(max_workers=len(groups)) as executor:
        futures = {
            executor.submit(_query_group, tag, rows, sessions[tag]): tag
            for tag, rows in groups.items()
        }
        for future in as_completed(futures):
            tag = futures[future]
            try:
                all_results.update(future.result())
            except Exception as e:
                _log(f"  [錯誤] {tag}: {e}")

    # 寫回 XLSX
    total = 0
    for row_num, (result, ts) in all_results.items():
        is_err = any(k in result for k in ["查不到", "錯誤", "失敗", "超過", "格式"])
        fill = err_fill if is_err else ok_fill

        rc = ws.cell(row=row_num, column=result_col, value=result)
        rc.font = font; rc.fill = fill
        rc.alignment = Alignment(vertical="center", wrap_text=True)

        tc = ws.cell(row=row_num, column=time_col, value=ts)
        tc.font = font; tc.fill = fill
        tc.alignment = Alignment(horizontal="center", vertical="center")
        total += 1

    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 20

    stem      = xlsx_path.rsplit(".", 1)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"{stem}_{timestamp}.xlsx"
    wb.save(out_path)
    print(f"\n完成，共查詢 {total} 筆，結果已儲存至 {out_path}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "查詢名單.xlsx"
    print(f"讀取 {path} ...")
    run_batch(path)
